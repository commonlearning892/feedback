#!/usr/bin/env python3
import argparse
import os
import re
import sys
from typing import Iterable, Tuple

try:
    from openpyxl import load_workbook
except Exception as e:
    print("ERROR: openpyxl is required. Install with: python3 -m pip install openpyxl", file=sys.stderr)
    raise

# Regexes to remove Tamil text and bracketed Tamil segments
RE_TAMIL_BLOCK = re.compile(r"[\u0B80-\u0BFF]+")
RE_PAREN_TAMIL_SEG = re.compile(r"\([^)]*[\u0B80-\u0BFF][^)]*\)")
RE_SPACES = re.compile(r"\s{2,}")
RE_EMPTY_PARENS = re.compile(r"\(\s*\)")
RE_TRAIL_PUNCT = re.compile(r"\s*[:;,-]\s*$")


def clean_text(text: str) -> str:
    if text is None:
        return text
    t = str(text)
    # Remove parenthesized segments that contain any Tamil characters
    t = RE_PAREN_TAMIL_SEG.sub("", t)
    # Remove any remaining Tamil characters
    t = RE_TAMIL_BLOCK.sub("", t)
    # Normalize dashes and whitespace, remove empty parens and trailing punctuation artifacts
    t = t.replace("–", "-").replace("—", "-")
    t = RE_EMPTY_PARENS.sub("", t)
    t = RE_SPACES.sub(" ", t).strip()
    t = RE_TRAIL_PUNCT.sub("", t)
    return t


def iter_xlsx_files(paths: Iterable[str]) -> Iterable[str]:
    for p in paths:
        if os.path.isdir(p):
            for root, _dirs, files in os.walk(p):
                for f in files:
                    if f.startswith("~$"):
                        continue
                    if f.lower().endswith(".xlsx") and not f.lower().endswith("_english.xlsx"):
                        yield os.path.join(root, f)
        else:
            if p.startswith("~$"):
                continue
            if p.lower().endswith(".xlsx") and not p.lower().endswith("_english.xlsx"):
                yield p


def process_file(path: str, inplace: bool = False, suffix: str = "_english") -> Tuple[str, int]:
    wb = load_workbook(path)
    changes = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                # Only attempt to clean strings
                if isinstance(v, str):
                    new_v = clean_text(v)
                    if new_v != v:
                        cell.value = new_v
                        changes += 1
    if inplace:
        out_path = path
    else:
        root, ext = os.path.splitext(path)
        out_path = f"{root}{suffix}{ext}"
    wb.save(out_path)
    return out_path, changes


def main():
    parser = argparse.ArgumentParser(description="Remove Tamil text from Excel files, keeping only plain English.")
    parser.add_argument("paths", nargs="*", default=["."], help="Files or directories to process (default: current directory)")
    parser.add_argument("--inplace", action="store_true", help="Modify the Excel files in place (destructive). By default, writes *_english.xlsx copies.")
    parser.add_argument("--suffix", default="_english", help="Suffix for output files when not using --inplace (default: _english)")
    args = parser.parse_args()

    files = list(iter_xlsx_files(args.paths))
    if not files:
        print("No .xlsx files found to process.")
        return 0

    total_changes = 0
    print(f"Found {len(files)} Excel file(s). Cleaning Tamil text...")
    for f in files:
        out_path, changes = process_file(f, inplace=args.inplace, suffix=args.suffix)
        total_changes += changes
        mode = "IN-PLACE" if args.inplace else f"COPY -> {os.path.basename(out_path)}"
        print(f"  - {os.path.basename(f)} => {mode} | cells changed: {changes}")

    print(f"Done. Total cells changed: {total_changes}")
    if not args.inplace:
        print("Note: Originals untouched. New *_english.xlsx files created alongside originals.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
